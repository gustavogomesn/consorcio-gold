from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum, Q
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl.workbook
from .models import Members, Loans, Meetings, TemporaryEntries, MeetingEntries, Fine, FundMovement
from .forms import UploadFileForm
from .utils import GenerateMeetingMinute
import openpyxl
import json
import io
import requests

LOAN_FEE = 3/100
MULT_ALLOWED_LOAN = 3
STOCK_VALUE = 50
FUND_VALUE = 5

def index(request):
    print('Server running')
    return JsonResponse({'Status': 'Server running'})

def get_meetings(request):
    meetings = list(Meetings.objects.order_by('id').values())
    return JsonResponse({'meetings': meetings})

@csrf_exempt
def new_meeting(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        meeting = Meetings(date = data['meeting-date'])
        meeting.save()
        
        # EVERY MEETING, THE FUND CONTRIBUTION IS REQUIRED TO ALL MEMBERS(member_id: 1 is only convencioned)
        fundContrib = FundMovement(**{
            'meeting': meeting,
            'member_id': 1,
            'value': 25*FUND_VALUE,
            'type': 'contribuition'
        })
        fundContrib.save()
        return JsonResponse({'status': 'Meeting registered'})
    return JsonResponse({'status': 'Method not allowed'})

@csrf_exempt
def delete_meeting(request):
    if request.method == "POST":
        data = json.loads(request.body)
        meeting = Meetings.objects.filter(id=data['id'])
        meeting.delete()
        return JsonResponse({'status': 'Meeting deleted'})
    return JsonResponse({'status': 'Method not allowed'})

def show_meeting(request, meeting_id):
    temporary_entries = list(TemporaryEntries.objects.filter(meeting_id=meeting_id, type='stocks').values('id', 'member__name', 'value'))
    temporary_payments = list(TemporaryEntries.objects.select_related('member').filter(meeting_id=meeting_id, type='loan-payment').values('id', 'member__name', 'value'))
    loans = list(Loans.objects.filter(meeting_id=meeting_id).values())
    fees_to_pay = list(Loans.objects.filter(isActive=True).filter(~Q(meeting_id=meeting_id)).values())
    fines_to_pay = list(Fine.objects.select_related('member').filter(paid=False).filter(~Q(meeting_id=meeting_id)).values('id', 'member__name', 'reason', 'value'))
    fund_movement = get_fund_movement(meeting_id)
    meeting_date = list(Meetings.objects.filter(id=meeting_id).values())[0]['date']
    return JsonResponse({'temporary_entries': temporary_entries,
                         'temporary_payments': temporary_payments,
                         'fees_to_pay': fees_to_pay,
                         'fines_to_pay': fines_to_pay,
                         'loans': loans,
                         'fund_movement': fund_movement,
                         'meeting_date': meeting_date})

def end_meeting(request, meeting_id):
    
    # MAKING FEE PAYMENTS
    fees_to_pay = Loans.objects.filter(isActive=True).filter(~Q(meeting_id=meeting_id)).values()
    print('## FEE COLECTED', fees_to_pay)
    meeting_entries_fees = [
        MeetingEntries(**{
            'type': 'fee',
            'value': entry['fee_by_month'],
            'meeting_id': meeting_id,
            'member_id': entry['member_id'],
            'loan_id': entry['id']
        })
    for entry in fees_to_pay
    ]
    MeetingEntries.objects.bulk_create(meeting_entries_fees)
    print('## FEE PAID')
    # MAKING FINE PAYMENTS
    fines_to_pay = Fine.objects.filter(paid=False).filter(~Q(meeting_id=meeting_id)).values()
    print('## FINES COLECTED', fines_to_pay)
    meeting_entries_fines = []
    for entry in fines_to_pay:
        # Populating entries
        meeting_entries_fines.append(
            MeetingEntries(**{
                'type': 'fine',
                'value': entry['value'],
                'meeting_id': meeting_id,
                'member_id': entry['member_id'],
            })
        )
        
        # Paiding fines
        fine = Fine.objects.get(id=entry['id'])
        fine.paid = True
        fine.save()
    MeetingEntries.objects.bulk_create(meeting_entries_fines)
    print('## FINES PAID')
    # MAKING CONTIBUITIONS(im setting columns in values() to avoid pass id)
    temporary_entries = TemporaryEntries.objects.filter(meeting_id=meeting_id).values('type', 'value', 'meeting_id', 'member_id', 'loan_id')
    meeting_entries = []
    for entry in temporary_entries:
        meeting_entries.append(MeetingEntries(**entry))
        print(MeetingEntries(**entry))
        if entry['type'] == 'stocks':
            member = Members.objects.get(id=entry['member_id'])
            member.stocks += entry['value']
            member.save()
    MeetingEntries.objects.bulk_create(meeting_entries)
    TemporaryEntries.objects.all().delete()
    print('## STOCKS PAID')
    # REGISTERING FUND MOVIMENTION
    fm = FundMovement.objects.filter(meeting_id=meeting_id).values()
    print('## FUND COLECTED', fm)
    fm_entries = [
        MeetingEntries(**{
            'type': 'fund_withdraw' if entry['type'] == '' else 'fund_contrib',
            'value': entry['value'],
            'meeting_id': meeting_id,
            'member_id': entry['member_id'],
        })
    for entry in fm
    ]
    for e in fm_entries:
        print(e)
    MeetingEntries.objects.bulk_create(fm_entries)
    print('## FUND PAID')
    # UPDATING NUMBERS OF MEETING AND ENDING IT
    sum_of_stocks = MeetingEntries.objects.filter(type='stocks', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    sum_of_loan_made = Loans.objects.filter(meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    sum_of_loan_payments = MeetingEntries.objects.filter(type='loan-payment', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    sum_of_fees = MeetingEntries.objects.filter(type='fee', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    sum_of_fines = MeetingEntries.objects.filter(type='fine', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    
    fund_contrib_value = MeetingEntries.objects.filter(type='fund_contrib', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    fund_withdraw_value = MeetingEntries.objects.filter(type='fund_withdraw', meeting_id=meeting_id).aggregate(Sum('value'))['value__sum']
    sum_of_fund = (fund_contrib_value or 0) - (fund_withdraw_value or 0)
    
    meeting = Meetings.objects.get(id=meeting_id)
    meeting.stocks = 0 if sum_of_stocks == None else sum_of_stocks
    meeting.loans_made = 0 if sum_of_loan_made == None else sum_of_loan_made
    meeting.loans_paid = 0 if sum_of_loan_payments == None else sum_of_loan_payments
    meeting.fees = 0 if sum_of_fees == None else sum_of_fees
    meeting.fines = 0 if sum_of_fines == None else sum_of_fines
    meeting.fund = 0 if sum_of_fund == None else sum_of_fund
    meeting.finished = True
    meeting.save()
    
    # UPDATING REMAING VALUE OF LOANS
    loan_payments = MeetingEntries.objects.filter(type='loan-payment', meeting_id=meeting_id).values()
    if len(loan_payments) != 0:
        for lp in loan_payments:
            loan = Loans.objects.get(id=lp['loan_id'])
            loan.remaining_value -= lp['value']
            loan.fee_by_month = loan.remaining_value * LOAN_FEE
            if loan.remaining_value == 0:
                loan.isActive = False
            loan.save()
            
    

    return JsonResponse({'status': 'Meeting finished'})

def members_model_download(request):
    with open ('website/assets/modelo_membros.xlsx', 'rb') as file:
        excel_file = file.read()

    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="modelo_membros.xlsx"'

    return response

def get_members(request):
    members = list(Members.objects.order_by('number').values())
    return JsonResponse({'members': members})

def handle_upload_members(file):
    try:
        wkb = openpyxl.load_workbook(file)
        sheet = wkb.active
        columns = sheet.max_column
        rows = sheet.max_row
        
        for r in range(1, rows):
            keys = ['number', 'name', 'role', 'contact', 'stocks']
            member = {
                'number': 0,
                'name': '',
                'role': '',
                'contact': '',
                'stocks': 0
            }
            for c in range(1, 6):
                member[keys[c-1]] = sheet.cell(r+1, c).value
            m = Members(**member)
            m.save()
    except:
        return JsonResponse({'status': 'This file isnt a spreadsheet'})
    
@csrf_exempt
def upload_members(request):
    if request.method == "POST":
        # i dont need this, after this will be removed i guess
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            handle_upload_members(request.FILES["file"])
            return JsonResponse({'status': 'File uploaded'})
    else:
        return JsonResponse({'status': 'Method isnt POST'})
    return JsonResponse({'status': 'Form invalid'})

@csrf_exempt
def new_loan(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        data['meeting'] = Meetings.objects.get(id=data['meeting_id'])   
        data['member'] = Members.objects.get(id = data['member_id'])
        data['value'] = data['loan-value']
        data['remaining_value'] = data['loan-value']
        data['fee_by_month'] = round(float(data['loan-value'])*LOAN_FEE, 2)
        del data['loan-value']
        print(data['meeting'].date)
        data['date'] = data['meeting'].date
        data['due_month'] = (data['date'] + relativedelta(months=+6)).strftime('%m-%Y')
        loan = Loans(**data)
        loan.save()
        return JsonResponse({'status': 'Loan registered'})

@csrf_exempt  
def delete_loan(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        loan = Loans.objects.filter(id = data['loan_id'])
        loan.delete()
        return JsonResponse({'status': 'Loan Deleted'})
    return JsonResponse({'status': 'Method not allowed'})

def get_loans(request):
    loans = list(Loans.objects.order_by('-date', '-isActive').values())
    return JsonResponse({'loans': loans})

@csrf_exempt  
def loan_payment_temporarily(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        print(data)
        meeting = Meetings.objects.get(id = data['meeting_id'])   
        loan = Loans.objects.get(id = data['loan_id'])
        loan_dict = Loans.objects.filter(id = data['loan_id']).values()[0]
        member = Members.objects.get(name = loan_dict['borrower'])
        loan = TemporaryEntries(meeting = meeting, loan = loan, member = member, type='loan-payment', value=data['payment'])
        loan.save()
        return JsonResponse({'status': 'Loan Deleted'})
    return JsonResponse({'status': 'Method not allowed'})

def contrib_model_download(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["numero", "nome", "acoes"])
    members = list(Members.objects.order_by('number').values())
    for member in members:
        ws.append([member['number'], member['name'], ''])

    # Save the workbook to an in-memory buffer
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="aportes.xlsx"'

    return response

def handle_upload_contribs(file, meeting_id):
    try:
        wkb = openpyxl.load_workbook(file)
        sheet = wkb.active
        columns = sheet.max_column
        rows = sheet.max_row
        for r in range(1, rows):
            keys = ['number', 'name', 'value']
            contrib = {
                'number': 0,
                'name': '',
                'value': '',
                'type': 'stocks',
                'meeting_id': meeting_id
            }
            for c in range(1, columns+1):
                contrib[keys[c-1]] = sheet.cell(r+1, c).value
            contrib['member_id'] = list(Members.objects.filter(number=contrib['number']).values())[0]['id']
            del contrib['number'], contrib['name']
            temp = TemporaryEntries(**contrib)
            temp.save()
        
        return JsonResponse({'status': 'Contribution uploaded successfully'})
            
    except:
        return JsonResponse({'status': 'This file isnt a spreadsheet'})
    
@csrf_exempt
def upload_contribs(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        handle_upload_contribs(request.FILES["file"], data['meeting_id'])
        return JsonResponse({'status': 'File uploaded'})
    else:
        return JsonResponse({'status': 'Method isnt POST'})
    
@csrf_exempt
def new_fine(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        fine = Fine(**{
                    'reason': data['reason'],
                    'value': data['fine-value'],
                    'member_id': data['member_id'],
                    'meeting_id': data['meeting_id']
                       })
        fine.save()
        return JsonResponse({'status': 'Fine registered'})

def minute_download(request, meeting_id):
    
    meeting = Meetings.objects.get(id=meeting_id)
    meeting_entries = MeetingEntries.objects.filter(meeting_id=meeting_id).values('type', 'value', 'member__name')
    loans = Loans.objects.filter(meeting_id=meeting_id).values('value', 'member__name', 'fee_by_month')
    fund = FundMovement.objects.filter(meeting_id=meeting_id).values('value', 'type', 'reason')
    summary = get_summary_data(meeting_id)

    pdf = GenerateMeetingMinute()
    generated_minute = io.BytesIO()
    pdf.generate(summary, meeting, meeting_entries, loans, fund,  generated_minute)
    generated_minute.seek(0)

    response = HttpResponse(generated_minute, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Ata_{meeting.date}.pdf"'

    return response

def get_summary(request, meeting_id=False):
    if not meeting_id:
        stocks = Meetings.objects.filter(finished = True).aggregate(Sum('stocks'))['stocks__sum']
        stocks_amount = Meetings.objects.filter(finished = True).aggregate(Sum('stocks'))['stocks__sum'] * STOCK_VALUE 
        fees = Meetings.objects.filter(finished = True).aggregate(Sum('fees'))['fees__sum']
        fines = Meetings.objects.filter(finished = True).aggregate(Sum('fines'))['fines__sum']
        loaned = Loans.objects.filter(isActive = True).aggregate(Sum('remaining_value'))['remaining_value__sum']
        cash = stocks_amount + fees + fines - loaned
        fund = Meetings.objects.filter(finished = True).aggregate(Sum('fund'))['fund__sum']
        current_stock_value = (cash + loaned) / stocks
        return JsonResponse({
            'stocks': stocks,
            'cash': cash,
            'loaned': loaned,
            'stock_value': current_stock_value,
            'fund': fund
        })
    else:
        stocks = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('stocks'))['stocks__sum']
        stocks_amount = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('stocks'))['stocks__sum'] * STOCK_VALUE
        fees = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fees'))['fees__sum']
        fines = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fines'))['fines__sum']
        loaned = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('loans_made'))['loans_made__sum'] - Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('loans_paid'))['loans_paid__sum']
        cash = stocks_amount + fees + fines - loaned
        current_stock_value = (cash + loaned) / stocks
        fund = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fund'))['fund__sum']
        return JsonResponse({
            'stocks': stocks,
            'cash': cash,
            'loaned': loaned,
            'stock_value': current_stock_value,
            'fund': fund
        })

def get_summary_data(meeting_id):
    stocks = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('stocks'))['stocks__sum']
    stocks_amount = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('stocks'))['stocks__sum'] * STOCK_VALUE
    fees = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fees'))['fees__sum']
    fines = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fines'))['fines__sum']
    loaned = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('loans_made'))['loans_made__sum'] - Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('loans_paid'))['loans_paid__sum']
    cash = stocks_amount + fees + fines - loaned
    current_stock_value = (cash + loaned) / stocks
    fund = Meetings.objects.filter(finished = True, id__lte = meeting_id).aggregate(Sum('fund'))['fund__sum']
    return ({
        'stocks': stocks,
        'cash': cash,
        'loaned': loaned,
        'stock_value': current_stock_value,
        'fund': fund
    })

# THIS ISNT A ROUTE, IS ONLY TO BE CONSUMED BY show_meeting ROUTE
def get_fund_movement(meeting_id):
    fm = list(FundMovement.objects.filter(meeting_id=meeting_id).values())
    return fm

@csrf_exempt
def fund_withdraw(request):
    if request.method == "POST":
        data = dict(request.POST.items())
        fund_movement = FundMovement(**{
                    'reason': data['reason'],
                    'value': data['fw-value'],
                    'member_id': data['member_id'],
                    'meeting_id': data['meeting_id']
                       })
        fund_movement.save()
        return JsonResponse({'status': 'Fund Withdraw registered'})
    return JsonResponse({'status': 'Method not allowed'})